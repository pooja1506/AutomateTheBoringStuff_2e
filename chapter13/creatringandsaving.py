import openpyxl
wb = openpyxl.Workbook() # Create a blank workbook.
wb.sheetnames # It starts with one sheet.
sheet = wb.active
sheet.title
sheet.title = 'Spam Bacon Eggs Sheet' # Change title.
wb.sheetnames