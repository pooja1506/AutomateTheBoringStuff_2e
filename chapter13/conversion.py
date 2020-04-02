import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
get_column_letter(1) # Translate column 1 to a letter.
get_column_letter(2)
get_column_letter(27)
get_column_letter(900)
wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
get_column_letter(sheet.max_column)
column_index_from_string('A') # Get A's number.
column_index_from_string('AA')


